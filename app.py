import openpyxl as xl

# Load workbook
wb = xl.load_workbook("data/sample.xlsx")
sheet = wb["Sheet1"]

# Apply 10% discount to prices in column C
for row in range(2, sheet.max_row + 1):
    price_cell = sheet.cell(row, 3)

    corrected_price = price_cell.value * 0.9

    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

# Save new workbook
wb.save("data/sample2.xlsx")

print("Excel automation completed successfully!")